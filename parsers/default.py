from openpyxl import load_workbook
from db import Group, Schedule, prepare_db
from datetime import datetime
import asyncio


class TableParser:
    def __init__(self, path_to_file):
        self.docs = load_workbook(path_to_file)
        self.table = self.docs.active
        self._merged_cells_fill()
        self.groups = []

    def _merged_cells_fill(self):
        for crange in self.table.merged_cells:
            for row in range(crange.min_row, crange.max_row + 1):
                for col in range(crange.min_col, crange.max_col + 1):
                    self.table.cell(row, col).value = self.table.cell(crange.min_row, crange.min_col).value

    async def parse(self):
        START_COL = 2
        START_ROW = 3
        LENGTH = 7
        WEEKS = 1
        self.groups = [self.table.cell(START_ROW, col).value for col in range(START_COL, START_COL + LENGTH)]
        for week in range(WEEKS):
            for day in range(6):
                for number, group in enumerate(self.groups):
                    date = str(self.table.cell(1 + 11 * week, 2 + LENGTH * day + number).value).split(" ")[0]
                    for p_n, row in enumerate(range(4 + 11 * week, 12 + 11 * week)):
                        if self.table.cell(row, 2 + LENGTH * day + number).value:
                            _group = await Group.get_or_create(group)
                            await Schedule.create(group_id=_group.id,
                                                  date=datetime.strptime(date, '%Y-%m-%d').date(),
                                                  pair_number=p_n + 1,
                                                  information=self.table.cell(row, 2 + LENGTH * day + number).value)


async def run():
    await prepare_db()

    parser = TableParser('../data/table.xlsx')
    await parser.parse()


if __name__ == '__main__':
    asyncio.run(run())
